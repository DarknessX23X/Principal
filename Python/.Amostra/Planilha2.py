import openpyxl
import os
import subprocess
import csv
import datetime
from openpyxl import load_workbook, formatting
import xlwings as xw

all = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='medium'),
    right=openpyxl.styles.Side(border_style='medium'),
    top=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='medium')
)

all_noright = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='thin'),
    top=openpyxl.styles.Side(border_style='thin'),
    bottom=openpyxl.styles.Side(border_style='thin')
)

bottom = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='medium')
)

bottom_thin = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='thin')
)

bottom_right = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='medium'),
    right=openpyxl.styles.Side(border_style='medium')
)
left_bottom = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='medium'),
    left=openpyxl.styles.Side(border_style='medium')
)
bottom_right_thin = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='medium'),
    right=openpyxl.styles.Side(border_style='medium')
)


workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = 'Receita'
sheet.sheet_view.showGridLines = False

sheet.merge_cells('A1:O1')
sheet['A1'] = 'Formulario de Receita'
sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=20)
for col in range(1, 16):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=1,column=col).border = all
sheet.row_dimensions[1].height = 35
sheet.merge_cells('A2:C2')
sheet['A2'] = 'Cliente:'
sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
for col in range(1, 4):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=2,column=col).border = all_noright
sheet.row_dimensions[2].height = 30
sheet.merge_cells('A3:C3')
sheet['A3'] = 'Lavagem:'
sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
for col in range(1, 4):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=3,column=col).border = all_noright
sheet.row_dimensions[3].height = 30
sheet['P1'] = 'Data:'
sheet['P1'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
sheet['P1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.cell(row=1,column=16).border = bottom
sheet.merge_cells('Q1:R1')
sheet.column_dimensions['Q'].width = 15
sheet.column_dimensions['R'].width = 15
for col in range(17, 18):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=1,column=col).border = bottom
sheet.cell(row=1,column=18).border = bottom_right
data = datetime.date.today()
data_formatada = data.strftime('%d/%m/%Y')
sheet['Q1'] = data_formatada
sheet['Q1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['Q1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000')
sheet.merge_cells('D2:L2')
sheet.merge_cells('D3:L3')
for col in range(4, 13):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=2,column=col).border = bottom_thin
    sheet.cell(row=3,column=col).border = bottom_thin
sheet.column_dimensions['M'].width = 10  
sheet['M2'] = 'Prod.:' 
sheet['M2'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['M2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.cell(row=2,column=13).border = left_bottom
sheet.cell(row=2,column=14).border = bottom_thin
sheet['M3'] = 'SPA:'
sheet['M3'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['M3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.cell(row=3,column=13).border = left_bottom
sheet.cell(row=2,column=14).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'), right=openpyxl.styles.Side(border_style='medium'))
sheet.merge_cells('N2:O2') 
sheet.cell(row=3,column=14).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'), right=openpyxl.styles.Side(border_style='medium'))
sheet.merge_cells('p2:q2') 
sheet['P2'] = 'Peso Unitário :'
sheet['P2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['p2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet['R2'] = '0,000'
sheet.cell(row=2,column=17).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=2,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(border_style='medium'))
sheet['R2'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
sheet['R2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000',italic=True)
sheet['O3'] = 'Qtde.:'
sheet.cell(row=3,column=15).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
sheet.cell(row=3,column=16).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(border_style='medium'), top=openpyxl.styles.Side(border_style='thin'))
sheet['O3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet['O3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['P3'] = '0'
sheet['P3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['P3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000',italic=True)
sheet['Q3'] = 'Peso T:'
sheet['Q3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['Q3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.cell(row=3,column=16).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=3,column=17).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
sheet.cell(row=3,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
sheet['R3'] = '0,000'
sheet['R3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['R3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.row_dimensions[4].height = 30
sheet['M4'] = 'Artigo:'
sheet['M4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['M4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.merge_cells('N4:R4')
sheet.cell(row=4,column=13).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'))
for col in range(14, 18):  # 1 = coluna A, 17 = coluna O
    sheet.cell(row=4,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
    sheet.cell(row=4,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
sheet.merge_cells('A4:E4')
sheet['A4'] = 'Antes de Lavar '
sheet['A4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)

#Coluna B
sheet.cell(row=5,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=7,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=9,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=11,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=13,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=15,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=17,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=19,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=21,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=23,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))


#Coluna D
sheet.cell(row=5,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=7,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=9,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=11,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=13,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))

#Coluna G
sheet.cell(row=5,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=7,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=9,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=11,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=13,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=15,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=17,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=19,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=21,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=23,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))



#Coluna J
sheet.cell(row=5,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=7,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=9,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=11,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=13,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=15,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=17,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=19,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
sheet.cell(row=21,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))


sheet.column_dimensions['A'].width = 1
sheet.column_dimensions['B'].width = 8
sheet.column_dimensions['C'].width = 9
sheet.column_dimensions['D'].width = 8
sheet.column_dimensions['E'].width = 9
#sheet.column_dimensions['F'].width = 1
sheet.column_dimensions['F'].hidden = True
sheet.column_dimensions['G'].width = 8
sheet.column_dimensions['H'].width = 10
sheet.column_dimensions['I'].width = 8
sheet.column_dimensions['J'].width = 10
sheet.column_dimensions['K'].width = 1
sheet.column_dimensions['L'].width = 1
sheet.column_dimensions['J'].width = 10
sheet.column_dimensions['Q'].width = 10
sheet.column_dimensions['R'].width = 10
sheet.row_dimensions[5].height = 30
sheet.row_dimensions[6].hidden =True
sheet.row_dimensions[7].height = 30
sheet.row_dimensions[8].hidden =True
sheet.row_dimensions[9].height = 30
sheet.row_dimensions[10].hidden =True
sheet.row_dimensions[11].height = 30
sheet.row_dimensions[12].hidden =True
sheet.row_dimensions[13].height = 30
sheet.row_dimensions[14].hidden =True
sheet.row_dimensions[15].height = 30
sheet.row_dimensions[16].hidden =True
sheet.row_dimensions[17].height = 30
sheet.row_dimensions[18].hidden =True
sheet.row_dimensions[19].height = 30
sheet.row_dimensions[20].hidden =True
sheet.row_dimensions[21].height = 30
sheet.row_dimensions[22].hidden =True
sheet.row_dimensions[23].height = 30
sheet.row_dimensions[24].height = 30
sheet.row_dimensions[25].height = 20

sheet['C5'] = 'Corrosão'
sheet['C5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['C5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
sheet['C7'] = 'Lixado'
sheet['C7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['C7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
sheet['C9'] = 'Esmeril'
sheet['C9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['C9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
sheet['C11'] = 'Amass.'
sheet['C11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['C11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
sheet['C13'] = ''
sheet['C13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['C13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)


sheet['E5'] = 'Used'
sheet['E5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['E5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['E7'] = 'Puido'
sheet['E7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['E7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['E9'] = 'Tanque'
sheet['E9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['E9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['E11'] = 'X TAG'
sheet['E11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['E11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['E13'] = 'Vinco'
sheet['E13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['E13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)

sheet['H5'] = 'Corrosão'
sheet['H5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['H5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['H7'] = 'Lixado'
sheet['H7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['H7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['H9'] = 'Esmeril'
sheet['H9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['H9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['H11'] = 'Amass.'
sheet['H11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['H11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['H13'] = ''
sheet['H13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['H13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)

sheet['J5'] = 'Used'
sheet['J5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['J7'] = 'Puido'
sheet['J7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['J9'] = 'Tanque'
sheet['J9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['J11'] = 'X TAG'
sheet['J11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['J13'] = 'Laser'
sheet['J13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['J15'] = 'Redinha'
sheet['J15'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['J15'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)


for row in range(4, 24):
    sheet.cell(row=row,column=6).border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'))
    
for col in range(1,13):
    sheet.cell(row=24,column=col).border = openpyxl.styles.Border(top=openpyxl.styles.Side(border_style='medium'))
for row in range(5, 24):
    sheet.cell(row=row,column=13).border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='medium'))

sheet.merge_cells('G4:L4') 
sheet['G4'] = 'Depois de Lavar:' 
sheet['G4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['G4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)   

sheet['M5'] = 'Tecido:' 
sheet['M5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
sheet['M5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)

sheet.merge_cells('M7:R7') 
sheet.merge_cells('M7:R7') 
sheet.merge_cells('M9:R9') 
sheet['M9'] = 'TAMANHO DA PEÇA:' 
sheet['M9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['M9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
sheet.merge_cells('M11:N11') 
sheet.merge_cells('O11:P11') 
sheet.merge_cells('Q11:R11')
sheet.merge_cells('M13:R13') 
sheet.merge_cells('A24:R24') 
sheet.merge_cells('A25:C25') 
sheet.merge_cells('D25:F25')
sheet.merge_cells('H25:N25')


sheet['M13'] = 'Lacres:' 
sheet['M13'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['M13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)

sheet['A24'] = 'Composição da Receita' 
sheet['A24'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A24'].font = openpyxl.styles.Font(name='Arial',bold=True, size=20)

sheet['A25'] = 'Processo' 
sheet['A25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['A25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['D25'] = '%' 
sheet['D25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['D25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['G25'] = 'SKU' 
sheet['G25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['G25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['H25'] = 'Produto' 
sheet['H25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['H25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['O25'] = 'Qtde' 
sheet['O25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['O25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['P25'] = 'Tempo' 
sheet['P25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['P25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['Q25'] = 'Temptra' 
sheet['Q25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['Q25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
sheet['R25'] = 'V.Banho' 
sheet['R25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['R25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)

for row in range(15,24,2):
    for col in range(13,19):
        sheet.cell(row=row,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

for col in range(13,19):
    sheet.cell(row=7,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
    sheet.cell(row=9,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))
    sheet.cell(row=11,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))
    sheet.cell(row=13,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

for col in range(1,19):
    sheet.cell(row=24,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

for col in range(1,19):
    sheet.cell(row=25,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

for row in range(26,63):
        sheet.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
        sheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=14)
        sheet.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)

for row in range(26,63):
        sheet.cell(row=row,column=1).alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet.cell(row=row,column=1).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
        sheet.cell(row=row,column=7).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
        sheet.cell(row=row,column=8).font=openpyxl.styles.Font(name='Arial',bold=True, size=8)


for row in range(26,63):
    for col in range(1,19):
        sheet.cell(row=row,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))

with open("C:\\CONFIG\\1.csv", 'r', newline='') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    # Lê o cabeçalho
    header = next(reader)
    row = 26
    for linha in reader:
     if linha[3] == "":
         sheet.cell(row=row, column=1).value = linha[7]
         row +=2

     else:
        sheet.cell(row=row, column=1).value = linha[7]
        sheet.cell(row=row, column=7).value = linha[3]
        sheet.cell(row=row, column=8).value = linha[4]
        row +=1
     #if linha[7] == "" :
                #sheet.cell(row=row, column=1).value = linha[2]               
                #row =row + 1       

if row > 62 :
    for row1 in range(63,row+1):
        sheet.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
        sheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=14)
        sheet.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)

    for row in range(63,row+1):
        sheet.cell(row=row,column=1).alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet.cell(row=row,column=1).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
        sheet.cell(row=row,column=7).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
        sheet.cell(row=row,column=8).font=openpyxl.styles.Font(name='Arial',bold=True, size=8)


    for row in range(63,row+1):
        for col in range(1,19):
            sheet.cell(row=row,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
    for row in range(63,row+1):
        sheet.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
        sheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=14)
        sheet.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)

workbook.save('E:\\Git\\Principal\\Python\\.Amostra\\Receita.xlsx')
os.startfile('E:\\Git\\Principal\\Python\\.Amostra\\Receita.xlsx')
#print('Arquivo Excel criado com sucesso!')