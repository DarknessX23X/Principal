from flask import Flask, render_template, request
import os
import openpyxl
import csv
import datetime
from openpyxl import load_workbook, formatting
import configparser



app = Flask(__name__)
current_dir = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
config.read(current_dir+'\\config\\config.ini')
secao = 'caminho' # Nome da seção
diretorio = config.get(secao, 'leitura')
salvar = config.get(secao, 'salvamento')
#exit()

#diretorio = 'X:\\TI\\Giovane\\Amostra'

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        arquivo_selecionado = request.form.get('arquivo')
        cliente = request.form.get('nome')
        arquivos = os.listdir(diretorio)
        arquivos_csv = [arquivo for arquivo in arquivos if arquivo.lower().endswith(('.csv', '.CSV'))]
        arquivos_sem_extensao = [os.path.splitext(arquivo)[0] for arquivo in arquivos_csv]
        arquivos_ordenados = sorted(arquivos_sem_extensao)
        nome = request.form.get('nome')
        data = request.form.get('data')
        peso_unitario = request.form.get('peso_unitario')
        lavagem = request.form.get('lavagem')
        spa = request.form.get('spa')
        transporte = request.form.get('Transporte')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Custo'
        sheet.merge_cells('B1:Q1')
        sheet.merge_cells('B2:B4')
        sheet.merge_cells('B5:C5')
        sheet.merge_cells('C2:L2')
        sheet.merge_cells('C3:L3')
        sheet.merge_cells('C4:L4')
        sheet.merge_cells('E5:L5')
        sheet.merge_cells('E6:L6')
        sheet.merge_cells('Q3:Q5')
        sheet['B1'] = 'CUSTO DE AMOSTRA'
        sheet['B2'] = 'Lacres>>'
        sheet['B5'] = 'Descrição>>>'
        sheet['B6'] = '%'
        sheet['C6'] = 'SKU.'
        sheet['E6'] = 'Produtos'
        sheet['Q2'] = 'Peso Unitário'
        sheet['Q3'] = f'{peso_unitario} Kg'
        sheet['Q6'] = 'VALOR R$'

        font = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=16)
        font2 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=12,underline='single')
        font3 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=10,underline='single')
        font4 = openpyxl.styles.Font(name='Arial',bold=True, size=11)
        font5 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=12)
        font6 = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        font7 = openpyxl.styles.Font(name='Arial',bold=True,  size=16)
        font8 = openpyxl.styles.Font(name='Arial',bold=True,  size=10,color='FF0000')
        font9 = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        font10 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=12)

        sheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet['C6'].alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet['Q3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['Q6'].alignment = openpyxl.styles.Alignment(horizontal='center')
        sheet['B1'].font = font
        sheet['B2'].font = font2
        sheet['B5'].font = font5
        sheet['B6'].font = font5
        sheet['C6'].font = font5
        sheet['E6'].font = font5
        sheet['Q2'].font = font3
        sheet['Q3'].font = font4
        sheet['Q6'].font = font5
        

        sheet.row_dimensions[1].height = 24.75
        sheet.row_dimensions[6].height = 18.75
        sheet.column_dimensions['B'].width = 10.14
        sheet.column_dimensions['C'].width = 9.14
        sheet.column_dimensions['Q'].width = 13.14
        sheet.column_dimensions['D'].hidden = True
        sheet.column_dimensions['E'].width = 6.86
        sheet.column_dimensions['F'].width = 6.86
        sheet.column_dimensions['G'].width = 6.86
        sheet.column_dimensions['H'].width = 6.86
        sheet.column_dimensions['I'].width = 6.86
        sheet.column_dimensions['J'].width = 6.86
        sheet.column_dimensions['K'].width = 6.86
        sheet.column_dimensions['L'].width = 6.86

        
        for col in range(13, 17):  
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True

        all = openpyxl.styles.Border(
        left=openpyxl.styles.Side(border_style='medium'),
        right=openpyxl.styles.Side(border_style='medium'),
        top=openpyxl.styles.Side(border_style='medium'),
        bottom=openpyxl.styles.Side(border_style='medium')
        )
        all_thin = openpyxl.styles.Border(
        left=openpyxl.styles.Side(border_style='thin'),
        right=openpyxl.styles.Side(border_style='thin'),
        top=openpyxl.styles.Side(border_style='medium'),
        bottom=openpyxl.styles.Side(border_style='thin')
        )

        all_noright = openpyxl.styles.Border(
        left=openpyxl.styles.Side(border_style='medium'),
        top=openpyxl.styles.Side(border_style='medium'),
        bottom=openpyxl.styles.Side(border_style='medium')
        )

        all_notop = openpyxl.styles.Border(
        left=openpyxl.styles.Side(border_style='medium'),
        right=openpyxl.styles.Side(border_style='medium'),
        bottom=openpyxl.styles.Side(border_style='medium')
        )

        all_bottom = openpyxl.styles.Border(
        bottom=openpyxl.styles.Side(border_style='medium')
        )

        all_right = openpyxl.styles.Border(
        right=openpyxl.styles.Side(border_style='medium')
        )   

        all_right_bottom = openpyxl.styles.Border(
        right=openpyxl.styles.Side(border_style='medium'),
        bottom=openpyxl.styles.Side(border_style='medium')    
        )   

        for row in range(1, 2):  
            for col in range(2, 18): 
                sheet.cell(row=row, column=col).border = all

        for row in range(2, 5):  
            sheet.cell(row=row,column=2).border = all_noright
            sheet.cell(row=row,column=17).border = all_right
            sheet.cell(row=row,column=12).border = all_right

        for column in range(3, 13):  
            sheet.cell(row=4,column=column).border = all_bottom

        for column in range(5, 13):  
            sheet.cell(row=5,column=column).border = all

        for column in range(2, 4):  
            sheet.cell(row=5,column=column).border = all

        for column in range(2, 18):  
            sheet.cell(row=6,column=column).border = all


        sheet.cell(row=4,column=12).border = all_right_bottom
        sheet.cell(row=5,column=17).border = all_notop
        sheet.sheet_view.showGridLines = False

        csv_file_path = f'{diretorio}\\{arquivo_selecionado}.CSV'
        with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
            class VariavelGlobal:
                def __init__(self, valor_inicial):
                    self.valor = valor_inicial

                def get_valor(self):
                    return self.valor

                def set_valor(self, novo_valor):
                    self.valor = novo_valor    
            subtotal = VariavelGlobal(0)        
            reader = csv.reader(csvfile, delimiter=';')
            next(reader) 
            total = 0
            row_num2 = 7
            for row in reader:
                if any(cell.strip() for cell in row):
                    if row[0] == "MP":  
                        DescricaoSDP = row[3] 
                        Sku = row[4]
                        DescricaoPRD = row[5]
                        Qtde = row[6] 
                        Valor = row[7]    
                        sheet.cell(row=5, column=5).value = DescricaoSDP
                        sheet.cell(row=5, column=5).alignment =openpyxl.styles.Alignment(horizontal='center') 
                        sheet.cell(row=row_num2, column=2).value = float(Qtde.replace(',', '.'))            
                        sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12)
                        sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center') 
                        sheet.cell(row=row_num2, column=3).value = int(Sku)
                        sheet.cell(row=row_num2, column=2).font = font6
                        sheet.cell(row=row_num2, column=3).font = font6
                        sheet.cell(row=row_num2, column=17).font = font6
                        sheet.cell(row=row_num2, column=3).alignment =openpyxl.styles.Alignment(horizontal='center')
                        sheet.cell(row=row_num2, column=5).value = DescricaoPRD 
                        sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='center')    
                        sheet.cell(row=row_num2, column=17).value = float(peso_unitario) * float(Valor.replace(',', '.')) * (float(Qtde.replace(',', '.'))) 
                        sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='center')
                        total = total + (float(peso_unitario) * float(Valor.replace(',', '.')) * (float(Qtde.replace(',', '.')))) 
                        subtotal.set_valor(total)
                        sheet.cell(row=row_num2, column=2).border = all_thin
                        sheet.cell(row=row_num2, column=3).border = all_thin            
                        sheet.cell(row=row_num2, column=17).border = all_thin

                        for col in range(5, 13):
                            sheet.cell(row=row_num2, column=col).border = all_thin
                        row_num2 += 1

                        for row_num3 in range(row_num2, row_num2 + 2):
                            for col in range(2, 13):
                                sheet.merge_cells(start_row=row_num3,end_row=row_num3,start_column=5,end_column=12)
                                sheet.cell(row=row_num3, column=col).border = all_thin
                        sheet.cell(row=row_num3, column=17).border = all_thin        
            row_num2 += 2       
            sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=12)    
            sheet.cell(row=row_num2, column=2).value = "MANUAIS / LAVAGEM"
            sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center')
            sheet.cell(row=row_num2, column=2).font = font
            for col in range(2, 18):
                sheet.cell(row=row_num2, column=col).border = all_thin
            row_num2 += 1
        with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
                class VariavelGlobal1:
                    def __init__(self, valor_inicial):
                        self.valor = valor_inicial

                    def get_valor(self):
                        return self.valor

                    def set_valor(self, novo_valor):
                        self.valor = novo_valor    
                subtotal2 = VariavelGlobal1(0)  
                total2 = 0     
                reader = csv.reader(csvfile, delimiter=';')
                next(reader)
                for row in reader:
                    if any(cell.strip() for cell in row):
                        if row[0] == "SV": 
                            sheet.cell(row=row_num2, column=3).value = int(row[4]) 
                            sheet.cell(row=row_num2, column=17).value = float(row[6].replace(',', '.'))
                            total2 = total2 + float(row[6].replace(',', '.'))
                            subtotal2.set_valor(total2) 
                            sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='center')
                            sheet.cell(row=row_num2, column=3).alignment =openpyxl.styles.Alignment(horizontal='center')
                            sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12)
                            sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='center') 
                            sheet.cell(row=row_num2, column=5).value = row[5]                 
                            for col in range(2, 18):
                                sheet.cell(row=row_num2, column=col).border = all_thin 
                            row_num2 += 1    
                            for row_num3 in range(row_num2, row_num2 + 3):
                                for col in range(2, 13):
                                    sheet.merge_cells(start_row=row_num3,end_row=row_num3,start_column=5,end_column=12)
                                    sheet.cell(row=row_num3, column=col).border = all_thin
                                    sheet.cell(row=row_num3, column=17).border = all_thin
                row_num2 += 2        
                sheet.row_dimensions[row_num2].height = 5.0
                for col in range(2,18):
                    sheet.cell(row=row_num2, column=col).border = all_thin
                row_num2 += 1
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
                data = datetime.date.today()
                data_formatada = data.strftime('%d/%m/%Y')
                sheet.row_dimensions[row_num2].height = 24
                sheet.cell(row=row_num2, column=2).value = data_formatada
                sheet.cell(row=row_num2, column=2).font = font6
                sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
                sheet.row_dimensions[row_num2].height = 24
                for col in range(2,18):
                    sheet.cell(row=row_num2, column=col).border = all_thin 
                    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                    sheet.cell(row=row_num2, column=5).value = "CUSTO PRODUTOS + 20%"  
                    sheet.cell(row=row_num2, column=17).value = 'R$ ' +  str(round(subtotal.get_valor() *1.2,4)).replace('.', ',')
                    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                    sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')  
                    sheet.cell(row=row_num2, column=5).font = font6 
                    sheet.cell(row=row_num2, column=17).font = font6      
                row_num2 += 1 
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
                sheet.cell(row=row_num2, column=2).value = "CALCULADO POR:"
                sheet.cell(row=row_num2, column=2).font = font6
                sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
                for col in range(2,18):
                    sheet.cell(row=row_num2, column=col).border = all_thin 
                    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                    sheet.cell(row=row_num2, column=5).value = "CUSTO MANUAIS/LAVAGEM"
                    sheet.cell(row=row_num2, column=17).value = 'R$ ' +  str(round(subtotal2.get_valor(),4)).replace('.', ',')
                    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
                    sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
                    sheet.cell(row=row_num2, column=5).font = font6
                    sheet.cell(row=row_num2, column=17).font = font6
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
                sheet.cell(row=row_num2, column=2).value = nome.upper()
                sheet.cell(row=row_num2, column=2).font = font10
                sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
                for col in range(2,18):
                    sheet.cell(row=row_num2, column=col).border = all_thin 
                    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                    sheet.cell(row=row_num2, column=5).value = "CUSTO SUB-TOTAL >>>"
                    sheet.cell(row=row_num2, column=17).value = 'R$ ' +  str(round((subtotal.get_valor()*1.2) + subtotal2.get_valor(),4)).replace('.', ',')
                    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                    sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')  
                    sheet.cell(row=row_num2, column=5).font = font6
                    sheet.cell(row=row_num2, column=17).font = font6
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2+4,start_column=2,end_column=3)
                sheet.cell(row=row_num2, column=2).value = "ATENÇÃO NOVOS PREÇOS "
                sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center',wrap_text=True) 
                sheet.cell(row=row_num2, column=2).font = font7
                for row1 in range(row_num2,row_num2+5):
                    for col in range(2,18):
                        sheet.cell(row=row1, column=col).border = all                
                        sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                        sheet.cell(row=row_num2, column=5).value = "TRANSPORTE"
                        sheet.cell(row=row_num2, column=17).value = "SIM"
                        sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                        sheet.cell(row=row_num2, column=5).font = font6   
                        sheet.cell(row=row_num2, column=17).font = font8        
                        sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                    for col in range(2,18):
                        sheet.cell(row=row1, column=col).border = all  
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                sheet.cell(row=row_num2, column=5).value = "COMISSÃO"
                sheet.cell(row=row_num2, column=17).value = "NÃO"
                sheet.cell(row=row_num2, column=5).font = font6
                sheet.cell(row=row_num2, column=17).font = font8  
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                for col in range(2,18):
                    sheet.cell(row=row1, column=col).border = all  
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                sheet.cell(row=row_num2, column=5).value = "TAMANHO DAS PEÇAS"
                sheet.cell(row=row_num2, column=17).value = "Pças Médias"
                sheet.cell(row=row_num2, column=5).font = font6
                sheet.cell(row=row_num2, column=17).font = font6
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
    
                for col in range(2,18):
                    sheet.cell(row=row1, column=col).border = all  
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                sheet.cell(row=row_num2, column=5).value = "PREÇO SUGERIDO PARA TERCEIROS (TOTAL + 20%)"
                sheet.cell(row=row_num2, column=17).value = 'R$ ' +  str(round((((subtotal.get_valor()*1.2) + subtotal2.get_valor() + float(transporte))*1.2),4)).replace('.', ',')
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
                sheet.cell(row=row_num2, column=5).font = font6
                sheet.cell(row=row_num2, column=17).font = font9
                for col in range(2,18):
                    sheet.cell(row=row1, column=col).border = all   
                row_num2 += 1
                sheet.row_dimensions[row_num2].height = 24
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
                sheet.cell(row=row_num2, column=5).value = "CUSTO TOTAL >>>"
                sheet.cell(row=row_num2, column=17).value = 'R$ ' +  str(round((subtotal.get_valor()*1.2) + subtotal2.get_valor() + float(transporte),4)).replace('.', ',')
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')  
                sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
                sheet.cell(row=row_num2, column=5).font = font7   
                sheet.cell(row=row_num2, column=17).font = font9           
                for col in range(2,18):
                    sheet.cell(row=row1, column=col).border = all  

        '''
        #Gera a Parte de Receita da Planilha        
        nova_planilha = workbook.create_sheet("Receitas", index=0)
        workbook.active = workbook['Receitas']
        sheet = workbook.active
        sheet['A1'] = "Dados da Receita" 
        all = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'),bottom=openpyxl.styles.Side(border_style='medium'))
        all_noright = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'), bottom=openpyxl.styles.Side(border_style='thin'))
        bottom = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
        bottom_thin = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'))
        bottom_right = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
        left_bottom = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'), left=openpyxl.styles.Side(border_style='medium'))
        bottom_right_thin = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells('A1:O1')
        sheet['A1'] = 'Formulario de Receita'
        sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=20)
        for col in range(1, 16):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=1,column=col).border = all
        sheet.row_dimensions[1].height = 35
        sheet.merge_cells('A2:C2')
        sheet['A2'] = 'Cliente:'
        sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        for col in range(1, 4):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=2,column=col).border = all_noright
        sheet.row_dimensions[2].height = 30
        sheet.merge_cells('A3:C3')
        sheet['A3'] = 'Lavagem:'
        sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        for col in range(1, 4):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=3,column=col).border = all_noright
        sheet.row_dimensions[3].height = 30
        sheet['P1'] = 'Data:'
        sheet['P1'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
        sheet['P1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.cell(row=1,column=16).border = bottom
        sheet.merge_cells('Q1:R1')
        sheet.column_dimensions['Q'].width = 15
        sheet.column_dimensions['R'].width = 15
        for col in range(17, 18):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=1,column=col).border = bottom
        sheet.cell(row=1,column=18).border = bottom_right
        data = datetime.date.today()
        data_formatada = data.strftime('%d/%m/%Y')
        sheet['Q1'] = data_formatada
        sheet['Q1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['Q1'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000')
        sheet.merge_cells('D2:L2')
        sheet.merge_cells('D3:L3')
        for col in range(4, 13):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=2,column=col).border = bottom_thin
            sheet.cell(row=3,column=col).border = bottom_thin
        sheet.column_dimensions['M'].width = 10  
        sheet['M2'] = 'Prod.:' 
        sheet['M2'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['M2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.cell(row=2,column=13).border = left_bottom
        sheet.cell(row=2,column=14).border = bottom_thin
        sheet['M3'] = 'SPA:'
        sheet['M3'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['M3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.cell(row=3,column=13).border = left_bottom
        sheet.cell(row=2,column=14).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'), right=openpyxl.styles.Side(border_style='medium'))
        sheet.merge_cells('N2:O2') 
        sheet.cell(row=3,column=14).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'), right=openpyxl.styles.Side(border_style='medium'))
        sheet.merge_cells('p2:q2') 
        sheet['P2'] = 'Peso Unitário :'
        sheet['P2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['p2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['R2'] = '0,000'
        sheet.cell(row=2,column=17).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=2,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(border_style='medium'))
        sheet['R2'].alignment = openpyxl.styles.Alignment(horizontal='right',vertical='center')
        sheet['R2'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000',italic=True)
        sheet['O3'] = 'Qtde.:'
        sheet.cell(row=3,column=15).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
        sheet.cell(row=3,column=16).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'), right=openpyxl.styles.Side(border_style='medium'), top=openpyxl.styles.Side(border_style='thin'))
        sheet['O3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['O3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['P3'] = '0'
        sheet['P3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['P3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14,color='FF0000',italic=True)
        sheet['Q3'] = 'Peso T:'
        sheet['Q3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['Q3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.cell(row=3,column=16).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=3,column=17).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
        sheet.cell(row=3,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
        sheet['R3'] = '0,000'
        sheet['R3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['R3'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.row_dimensions[4].height = 30
        sheet['M4'] = 'Artigo:'
        sheet['M4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['M4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.merge_cells('N4:R4')
        sheet.cell(row=4,column=13).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'))
        for col in range(14, 18):  # 1 = coluna A, 17 = coluna O
            sheet.cell(row=4,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'))
            sheet.cell(row=4,column=18).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
        sheet.merge_cells('A4:E4')
        sheet['A4'] = 'Antes de Lavar '
        sheet['A4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.cell(row=5,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=7,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=9,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=11,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=13,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=15,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=17,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=19,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=21,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=23,column=2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=5,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=7,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=9,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=11,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=13,column=4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=5,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=7,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=9,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=11,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=13,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=15,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=17,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=19,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=21,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=23,column=7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=5,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=7,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=9,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=11,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=13,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=15,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=17,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=19,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.cell(row=21,column=9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        sheet.column_dimensions['A'].width = 1
        sheet.column_dimensions['B'].width = 8
        sheet.column_dimensions['C'].width = 9
        sheet.column_dimensions['D'].width = 8
        sheet.column_dimensions['E'].width = 9
        #sheet.column_dimensions['F'].width = 1
        sheet.column_dimensions['F'].hidden = True
        sheet.column_dimensions['G'].width = 8
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 8
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 1
        sheet.column_dimensions['L'].width = 1
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['Q'].width = 10
        sheet.column_dimensions['R'].width = 10
        sheet.row_dimensions[5].height = 30
        sheet.row_dimensions[6].hidden =True
        sheet.row_dimensions[7].height = 30
        sheet.row_dimensions[8].hidden =True
        sheet.row_dimensions[9].height = 30
        sheet.row_dimensions[10].hidden =True
        sheet.row_dimensions[11].height = 30
        sheet.row_dimensions[12].hidden =True
        sheet.row_dimensions[13].height = 30
        sheet.row_dimensions[14].hidden =True
        sheet.row_dimensions[15].height = 30
        sheet.row_dimensions[16].hidden =True
        sheet.row_dimensions[17].height = 30
        sheet.row_dimensions[18].hidden =True
        sheet.row_dimensions[19].height = 30
        sheet.row_dimensions[20].hidden =True
        sheet.row_dimensions[21].height = 30
        sheet.row_dimensions[22].hidden =True
        sheet.row_dimensions[23].height = 30
        sheet.row_dimensions[24].height = 30
        sheet.row_dimensions[25].height = 20
        sheet['C5'] = 'Corrosão'
        sheet['C5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['C5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        sheet['C7'] = 'Lixado'
        sheet['C7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['C7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        sheet['C9'] = 'Esmeril'
        sheet['C9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['C9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        sheet['C11'] = 'Amass.'
        sheet['C11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['C11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        sheet['C13'] = ''
        sheet['C13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['C13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=10)
        sheet['E5'] = 'Used'
        sheet['E5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['E5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['E7'] = 'Puido'
        sheet['E7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['E7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['E9'] = 'Tanque'
        sheet['E9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['E9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['E11'] = 'X TAG'
        sheet['E11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['E11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['E13'] = 'Vinco'
        sheet['E13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['E13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H5'] = 'Corrosão'
        sheet['H5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['H5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H7'] = 'Lixado'
        sheet['H7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['H7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H9'] = 'Esmeril'
        sheet['H9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['H9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H11'] = 'Amass.'
        sheet['H11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['H11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H13'] = ''
        sheet['H13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['H13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J5'] = 'Used'
        sheet['J5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J7'] = 'Puido'
        sheet['J7'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J7'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J9'] = 'Tanque'
        sheet['J9'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J11'] = 'X TAG'
        sheet['J11'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J11'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J13'] = 'Laser'
        sheet['J13'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['J15'] = 'Redinha'
        sheet['J15'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['J15'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        for row in range(4, 24):
            sheet.cell(row=row,column=6).border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin'))    
        for col in range(1,13):
            sheet.cell(row=24,column=col).border = openpyxl.styles.Border(top=openpyxl.styles.Side(border_style='medium'))
        for row in range(5, 24):
            sheet.cell(row=row,column=13).border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='medium'))
        sheet.merge_cells('G4:L4') 
        sheet['G4'] = 'Depois de Lavar:' 
        sheet['G4'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['G4'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['M5'] = 'Tecido:' 
        sheet['M5'].alignment = openpyxl.styles.Alignment(horizontal='left',vertical='center')
        sheet['M5'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.merge_cells('M7:R7') 
        sheet.merge_cells('M7:R7') 
        sheet.merge_cells('M9:R9') 
        sheet['M9'] = 'TAMANHO DA PEÇA:' 
        sheet['M9'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['M9'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet.merge_cells('M11:N11') 
        sheet.merge_cells('O11:P11') 
        sheet.merge_cells('Q11:R11')
        sheet.merge_cells('M13:R13') 
        sheet.merge_cells('A24:R24') 
        sheet.merge_cells('A25:C25') 
        sheet.merge_cells('D25:F25')
        sheet.merge_cells('H25:N25')
        sheet['M13'] = 'Lacres:' 
        sheet['M13'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['M13'].font = openpyxl.styles.Font(name='Arial',bold=True, size=14)
        sheet['A24'] = 'Composição da Receita' 
        sheet['A24'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A24'].font = openpyxl.styles.Font(name='Arial',bold=True, size=20)
        sheet['A25'] = 'Processo' 
        sheet['A25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['A25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['D25'] = '%' 
        sheet['D25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['D25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['G25'] = 'SKU' 
        sheet['G25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['G25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['H25'] = 'Produto' 
        sheet['H25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['H25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['O25'] = 'Qtde' 
        sheet['O25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['O25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['P25'] = 'Tempo' 
        sheet['P25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['P25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['Q25'] = 'Temptra' 
        sheet['Q25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['Q25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)
        sheet['R25'] = 'V.Banho' 
        sheet['R25'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet['R25'].font = openpyxl.styles.Font(name='Arial',bold=True, size=12)

        for row in range(15,24,2):
            for col in range(13,19):
                sheet.cell(row=row,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

        for col in range(13,19):
            sheet.cell(row=7,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'))
            sheet.cell(row=9,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))
            sheet.cell(row=11,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))
            sheet.cell(row=13,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

        for col in range(1,19):
            sheet.cell(row=24,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

        for col in range(1,19):
            sheet.cell(row=25,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='medium'),left=openpyxl.styles.Side(border_style='medium'),right=openpyxl.styles.Side(border_style='medium'),top=openpyxl.styles.Side(border_style='medium'))

        for row in range(26,63):
            sheet.merge_cells(start_row=row,start_column=1,end_row=row,end_column=3)
            sheet.merge_cells(start_row=row,start_column=8,end_row=row,end_column=14)
            sheet.merge_cells(start_row=row,start_column=4,end_row=row,end_column=5)

        for row in range(26,63):
            sheet.cell(row=row,column=1).alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
            sheet.cell(row=row,column=1).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
            sheet.cell(row=row,column=7).font=openpyxl.styles.Font(name='Arial',bold=True, size=8,color='FF0000',italic=True)
            sheet.cell(row=row,column=8).font=openpyxl.styles.Font(name='Arial',bold=True, size=8)


        for row in range(26,63):
            for col in range(1,19):
                sheet.cell(row=row,column=col).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(border_style='thin'),left=openpyxl.styles.Side(border_style='thin'),right=openpyxl.styles.Side(border_style='thin'),top=openpyxl.styles.Side(border_style='thin'))
        '''


        workbook.save(f'{salvar}\\{arquivo_selecionado}.xlsx')
        return render_template('index.html', arquivo_selecionado=arquivo_selecionado, arquivos=arquivos_ordenados)
    else:
        arquivos = os.listdir(diretorio)
        arquivos_csv = [arquivo for arquivo in arquivos if arquivo.lower().endswith(('.csv', '.CSV'))]
        arquivos_sem_extensao = [os.path.splitext(arquivo)[0] for arquivo in arquivos_csv]
        arquivos_ordenados = sorted(arquivos_sem_extensao)
        return render_template('index.html', arquivos=arquivos_ordenados)

if __name__ == '__main__':
    app.run(debug=True)