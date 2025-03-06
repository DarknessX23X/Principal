import openpyxl
from openpyxl.styles import PatternFill
from ldap3 import Server, Connection, ALL_ATTRIBUTES, ALL

# Conectando ao domínio AD
server = Server('192.168.0.53', get_info=ALL)
ldap_user = 'administrador@kazzo1.sys'
ldap_password = 'Tec2024K@3,14colli'

# Conectando ao servidor AD
conn = Connection(server, user=ldap_user, password=ldap_password)
conn.bind()

# Consulta ao computador no AD
base_dn = 'cn=Computers,dc=kazzo1,dc=sys'  # Distinguished Name do container "Computers"
computers_filter = '(objectClass=computer)'
conn.search(base_dn, computers_filter, attributes=ALL_ATTRIBUTES)
computers = conn.entries

# Ordena os resultados pelo nome do computador em ordem crescente
computers.sort(key=lambda x: x.cn.value)

# Cria um arquivo XLSX
wb = openpyxl.Workbook()
ws = wb.active

# Define o cabeçalho das colunas
header = ['Nome do Computador', 'Descrição', 'Sistema Operacional']

# Escreve o cabeçalho na primeira linha
for col_num, col_title in enumerate(header, 1):
    ws.cell(row=1, column=col_num, value=col_title)

# Percorre os resultados e escreve cada linha
for row_num, computer_entry in enumerate(computers, 2):
    computer_name = computer_entry.cn.value
    description = computer_entry.description.value if hasattr(computer_entry, 'description') else ""
    operating_system = computer_entry.operatingSystem.value if hasattr(computer_entry, 'operatingSystem') else ""

    # Escreve os valores nas células
    ws.cell(row=row_num, column=1, value=computer_name)
    ws.cell(row=row_num, column=2, value=description)
    ws.cell(row=row_num, column=3, value=operating_system)

    # Verifica se a descrição está em branco e pinta a célula correspondente de vermelho
    if not description:
        fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        ws.cell(row=row_num, column=2).fill = fill

    # Verifica se o sistema operacional é 'Windows 7 Professional' e pinta a célula correspondente de amarelo
    if operating_system == 'Windows 7 Professional':
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws.cell(row=row_num, column=3).fill = fill

# Salva o arquivo XLSX
wb.save('computadores.xlsx')

# Fechando a conexão
conn.unbind()