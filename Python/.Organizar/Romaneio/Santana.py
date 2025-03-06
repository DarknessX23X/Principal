import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# Ler o arquivo HTML e selecionar o DataFrame desejado
dfs = pd.read_html('.\\Romaneio\\Romaneio.htm')
df = dfs[1]
df.to_excel('.\\Romaneio\\Romaneio.xlsx', index=False, header=False)

workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]

df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:].astype({df.columns[13]: float}) 
df[df.columns[13]] = df[df.columns[13]] / 100 
for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value) 

workbook.save('.\\Romaneio\\Romaneio.xlsx')


workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]

df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:].astype({df.columns[12]: float}) 
df[df.columns[12]] = df[df.columns[12]] / 100 
for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value) 

workbook.save('.\\Romaneio\\Romaneio.xlsx')


workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]

df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:].astype({df.columns[9]: float}) 
df[df.columns[9]] = df[df.columns[9]] / 100 
for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value) 

workbook.save('.\\Romaneio\\Romaneio.xlsx')

workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]

df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:].astype({df.columns[14]: float}) 
df[df.columns[14]] = df[df.columns[14]] / 100 
for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value) 

workbook.save('.\\Romaneio\\Romaneio.xlsx')

workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
sheet_name = workbook.sheetnames[0]
worksheet = workbook[sheet_name]
df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:]

df[df.columns[11]] = df[df.columns[11]].astype(str).str.lstrip('0')
df[df.columns[2]] = pd.to_numeric(df[df.columns[2]], errors='coerce')
df[df.columns[3]] = pd.to_numeric(df[df.columns[3]], errors='coerce')
df[df.columns[4]] = pd.to_numeric(df[df.columns[4]], errors='coerce')
df[df.columns[5]] = pd.to_numeric(df[df.columns[5]], errors='coerce')
df[df.columns[6]] = pd.to_numeric(df[df.columns[6]], errors='coerce')
df[df.columns[7]] = pd.to_numeric(df[df.columns[7]], errors='coerce')
df[df.columns[8]] = pd.to_numeric(df[df.columns[8]], errors='coerce')
df[df.columns[10]] = pd.to_numeric(df[df.columns[10]], errors='coerce')
df[df.columns[11]] = pd.to_numeric(df[df.columns[11]], errors='coerce')


for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value) 
workbook.save('.\\Romaneio\\Romaneio.xlsx')



exit()

workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
worksheet = workbook['Sheet1']
worksheet.delete_cols(1, 4)
new_column_index = 1
for column in worksheet.iter_cols(min_col=1, max_col=worksheet.max_column):
    column = [cell for cell in column]
    for cell in column:
        worksheet.cell(row=cell.row, column=new_column_index, value=cell.value)
    new_column_index += 1
workbook.save('.\\Romaneio\\Romaneio.xlsx')

workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')
worksheet = workbook['Sheet1']
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=3):
    value_c = row[2].value  # Valor da coluna C
    value_a = row[0].value  # Valor da coluna A

    row[0].value = value_c  # Mover o valor da coluna C para a coluna A
    row[2].value = value_a  # Mover o valor da coluna A para a coluna C

workbook.save('.\\Romaneio\\Romaneio.xlsx')
workbook = openpyxl.load_workbook('.\\Romaneio\\Romaneio.xlsx')

# Acessar a planilha desejada
worksheet = workbook['Sheet1']

# Converter o DataFrame do pandas
df = pd.DataFrame(worksheet.values)
df.columns = df.iloc[0]
df = df[1:].astype({df.columns[8]: float})  # Substitua 8 pelo índice numérico da coluna I

# Dividir os valores na coluna I por 100
df[df.columns[8]] = df[df.columns[8]] / 100  # Substitua 8 pelo índice numérico da coluna I

# Atualizar os valores na planilha
for index, row in df.iterrows():
    for col_index, value in enumerate(row):
        worksheet.cell(row=index+1, column=col_index+1, value=value)  # Adicionar células diretamente na planilha

# Salvar o arquivo Excel atualizado
workbook.save('.\\Romaneio\\Romaneio.xlsx')