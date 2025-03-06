import pyodbc
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table
from openpyxl.utils import range_boundaries
from datetime import datetime
import os
import time

mdb_file = r'\\192.168.0.167\\c$\\Audaces Ultraspool 8\\pltLog.mdb'
conn_str = r'Driver={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={};PWD={};'.format(mdb_file, "sinajet")
excel_file = r'X:\\TI\\.XLSX\\Plotados.xlsx'

#print(current_datetime)


while True:
    current_minute = datetime.now().minute
    if current_minute % 10 == 0 or current_minute % 10 == 5:
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_ren = r'X:\\TI\\.XLSX\\Plotados_{}.xlsx'.format(current_datetime)
        try:
            os.rename(excel_file, excel_ren)
        except Exception as e:
            print("Ocorreu um erro ao renomear o arquivo:", e)
        
        conn = pyodbc.connect(conn_str)
        cursor = conn.cursor()
        sql_query = 'SELECT * FROM PlotHistory'
        df = pd.read_sql_query(sql_query, conn)        
        workbook = Workbook()
        worksheet = workbook.active
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        data_alignment = Alignment(horizontal='center')
        zebra_fill = PatternFill(fill_type="solid", fgColor="E0E0E0")  # Cor de fundo cinza
        current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")

# Preencha a tabela com cabeçalhos de coluna
        for col_num, column_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=1, column=col_num, value=column_name)
            cell.font = header_font
            cell.alignment = header_alignment


# Preencha a tabela com dados e aplique estilos zebrados
        for row_num, row_data in enumerate(df.values, start=2):
            for col_num, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = data_alignment
                if row_num % 2 == 0:  # Aplica estilo zebrado para linhas pares
                    cell.fill = zebra_fill

# Ajuste a largura das colunas com base no tamanho da maior palavra de cada coluna
        for col_num, column_name in enumerate(df.columns, start=1):
            max_length = max(df[column_name].astype(str).map(len).max(), len(column_name))
            adjusted_width = (max_length + 2) * 1.2  # Fator de ajuste
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = adjusted_width

            if "Plotados" in worksheet._tables:
                del worksheet._tables["Plotados"]

            table_name = "Plotados"    
    
            table_range = f"A1:{worksheet.cell(row=len(df)+1, column=len(df.columns)).coordinate}"            
            table_obj = Table(displayName="Plotados", ref=table_range)
            worksheet.add_table(table_obj)
            


# Salve o workbook no arquivo Excel
        workbook.save(excel_file)

# Feche a conexão
        conn.close()

    time.sleep(60)