from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from openpyxl import load_workbook
import time
import datetime
import locale
import subprocess
from ping3 import ping

setor = []
contador =[]
locale.setlocale(locale.LC_TIME, 'pt_BR.utf8')

chrome_options = Options()
chrome_options.add_argument("--headless") 
driver = webdriver.Chrome(options=chrome_options)

def check_ping(hostname):
    response = ping(hostname)
    if response is not None:
        print(f"{hostname} está online. Tempo de resposta: {response} segundos.")
    else:
        print(f"{hostname} está offline ou inacessível.")

# Exemplo de uso
        
#sites = ['192.168.0.60', '192.168.0.61', '192.168.0.62','192.168.0.63','192.168.0.64','192.168.0.65','192.168.0.64','192.168.0.65','192.168.0.67','192.168.0.68','192.168.0.69','192.168.0.70','192.168.10.55']
#for site in sites:
#    check_ping(site)
#exit()



 #Almoxarifado
response = ping('192.168.0.60')
if response is not None:
    driver.get("http://192.168.0.60/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[15:30])
    setor.append("Almoxarifado")
    contador.append(elemento[4].text[15:30])

 #Etiquetas
response = ping('192.168.0.61')
if response is not None:    
    driver.get("http://192.168.0.61/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[15:30])
    setor.append("Etiquetas")
    contador.append(elemento[4].text[15:30])

#Financeiro
response = ping('192.168.0.62')
if response is not None:     
    driver.get("http://192.168.0.62/general/information.html?kind=item")
    elemento = driver.find_element("name","B134d")
    elemento.send_keys("4&05fpPB")
    elemento.send_keys(Keys.RETURN)
    elemento = driver.find_elements(By.TAG_NAME,"dd")
    print(elemento[5].text)
    setor.append("Financeiro")
    contador.append(elemento[5].text)

#Colorida
response = ping('192.168.0.63')
if response is not None:  
    driver.get("http://192.168.0.63/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[5].text[8:30])
    print(elemento[6].text[6:30])
    setor.append("Colorida")
    contador.append(elemento[5].text[8:30])
    setor.append("Colorida Preto")
    contador.append(elemento[6].text[6:30])

 #Administrativo
response = ping('192.168.0.64')
if response is not None:
    driver.get("http://192.168.0.64/general/information.html?kind=item")
    elemento = driver.find_element("name","B1350")
    elemento.send_keys("initpass")
    elemento.send_keys(Keys.RETURN)
    elemento = driver.find_elements(By.TAG_NAME,"dd")
    print(elemento[5].text)
    setor.append("Administrativo")
    contador.append(elemento[5].text)

 #Departamento Pessoal
response = ping('192.168.0.65')
if response is not None:
    driver.get("http://192.168.0.65/general/information.html?kind=item")
    elemento = driver.find_element("name","B134d")
    elemento.send_keys("initpass")
    elemento.send_keys(Keys.RETURN)
    elemento = driver.find_elements(By.TAG_NAME,"dd")
    print(elemento[5].text)
    setor.append("Departamento Pessoal")
    contador.append(elemento[5].text)

 #Recebimento
response = ping('192.168.0.67')
if response is not None:
    driver.get("http://192.168.0.67/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[15:30])
    setor.append("Recebimento")
    contador.append(elemento[4].text[15:30])

 #Lavanderia
response = ping('192.168.0.68')
if response is not None:
    driver.get("http://192.168.0.68/printer/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[13:30])
    setor.append("Lavanderia")
    contador.append(elemento[4].text[13:30])


#Expedição
response = ping('192.168.0.69')
if response is not None:
    driver.get("http://192.168.0.69/general/information.html?kind=item")
    elemento = driver.find_element("name","B134e")
    elemento.send_keys("initpass")
    elemento.send_keys(Keys.RETURN)
    elemento = driver.find_elements(By.TAG_NAME,"dd")
    print(elemento[5].text)
    setor.append("Expedição")
    contador.append(elemento[5].text)

 #Modelagem
response = ping('192.168.0.70')
if response is not None:    
    driver.get("http://192.168.0.70/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[15:30])
    setor.append("Modelagem")
    contador.append(elemento[4].text[15:30])

#Posto Trevao
response = ping('192.168.10.55')
if response is not None:     
    driver.get("http://192.168.10.55/main/main.html")
    elemento = driver.find_elements(By.CLASS_NAME,"elemwhite")
    print(elemento[4].text[15:30])
    setor.append("Trevao")
    contador.append(elemento[4].text[15:30])


print(setor)
print(contador)
workbook = Workbook()

workbook = load_workbook(filename="C:\\config\\TI - Infobras.xlsx")

sheet = workbook.active

sheet['E3'] = str(sheet.cell(row=3, column=6).value)
sheet['E4'] = int(sheet.cell(row=4, column=6).value)
sheet['E5'] = int(sheet.cell(row=5, column=6).value)
sheet['E6'] = int(sheet.cell(row=6, column=6).value)
sheet['E7'] = int(sheet.cell(row=7, column=6).value)
sheet['E8'] = int(sheet.cell(row=8, column=6).value)
sheet['E9'] = int(sheet.cell(row=9, column=6).value)
sheet['E10'] = int(sheet.cell(row=10, column=6).value)
sheet['E11'] = int(sheet.cell(row=11, column=6).value)
sheet['E12'] = int(sheet.cell(row=12, column=6).value)
sheet['E13'] = int(sheet.cell(row=13, column=6).value)

data_atual = datetime.date.today()
primeiro_dia_mes_atual = data_atual.replace(day=1)
ultimo_dia_mes_anterior = primeiro_dia_mes_atual - datetime.timedelta(days=1)

sheet['F3'] = ultimo_dia_mes_anterior.strftime('%B').title()
sheet['F4'] = int(contador[6]) #Departamento Pessoal
sheet['F5'] = int(contador[5]) #Administrativo
sheet['F6'] = int(contador[2]) #Financeiro
sheet['F7'] = int(contador[0]) #Almoxarifado
sheet['F8'] = int(contador[1]) #Etiquetas
sheet['F9'] = int(contador[9]) #Expedicao
sheet['F10'] = int(contador[8])#Lavanderia
sheet['F11'] = int(contador[7]) #Recebimento
sheet['F12'] = int(contador[11]) #Trevao
sheet['F13'] = '0' #Kabmadeira
sheet['F14'] = int(contador[4])#Colorida Preto
sheet['F15'] = int(contador[3])#Colorida
sheet['F16'] = int(contador[10])#Modelagem



workbook.save("C:\\config\\TI - Infobras.xlsx")


