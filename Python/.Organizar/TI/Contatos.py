
from openpyxl import load_workbook
import csv
# Carregar o arquivo Excel (.xlsx)

header = [
    'Name', 'Number', 'First Name', 'Last Name', 'Phone Number', 'Mobile Number',
    'E-mail Address', 'Address', 'City', 'State', 'Postal Code', 'Comment',
    'Id', 'Info', 'Presence', 'Directory', 'Starred'
]
additional_info = ',,,,,,,,,,,,,0,0,0'
workbook = load_workbook(filename='X:\\TI\\Giovane\\ramais_final.xlsx')

# Selecionar a primeira planilha no arquivo
sheet = workbook.active

with open('colunas_selecionadas.csv', mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(header)  # Escrever o cabeçalho no arquivo CSV
    for row in sheet.iter_rows(values_only=True):
        writer.writerow([row[1], row[0], additional_info])