from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import Workbook


chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(chrome_options)
driver.get("http://192.168.0.1:14000/")
elemento = driver.find_element("name","usernamefld")
elemento.send_keys("giovane.adm")
elemento = driver.find_element("name","passwordfld")
elemento.send_keys("DarknessX23X")
elemento.send_keys(Keys.RETURN)
driver.get("http://192.168.0.1:14000/services_dhcp.php")
elemento = driver.find_elements(By.TAG_NAME,"td")

valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}

workbook = Workbook()

sheet = workbook.active

sheet.cell(row=1, column=1).value = 'MAC'
sheet.cell(row=1, column=2).value = 'NOME'
sheet.cell(row=1, column=3).value = 'IP'
sheet.cell(row=1, column=4).value = 'HOSTNAME'
sheet.cell(row=1, column=5).value = 'DESCRIÇÃO'

#print(range(len(sublistas)))

#exit()
sheet.title ='FIREWALL'

for i in range(len(sublistas)):
    sheet.cell(row=i+2, column=1).value = str(sublistas[i][0])
    sheet.cell(row=i+2, column=2).value = str(sublistas[i][1])
    sheet.cell(row=i+2, column=3).value = str(sublistas[i][2])
    sheet.cell(row=i+2, column=4).value = str(sublistas[i][3])
    sheet.cell(row=i+2, column=5).value = str(sublistas[i][4])


driver.get("http://192.168.0.27:14000/")
elemento = driver.find_element("name","usernamefld")
elemento.send_keys("giovane")
elemento = driver.find_element("name","passwordfld")
elemento.send_keys("DarknessX23X")
elemento.send_keys(Keys.RETURN)
driver.get("http://192.168.0.27:14000/services_dhcp.php")
elemento = driver.find_elements(By.TAG_NAME,"td")

valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}

sheet2 = workbook.create_sheet('ROTEADOR')
sheet2.cell(row=1, column=1).value = 'MAC'
sheet2.cell(row=1, column=2).value = 'NOME'
sheet2.cell(row=1, column=3).value = 'IP'
sheet2.cell(row=1, column=4).value = 'HOSTNAME'
sheet2.cell(row=1, column=5).value = 'DESCRIÇÃO'

for i in range(len(sublistas)):
    sheet2.cell(row=i+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=i+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=i+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=i+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=i+2, column=5).value = str(sublistas[i][4])
    
x=i
print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt1")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt2")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt3")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt4")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt5")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt6")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

driver.get("http://192.168.0.27:14000/services_dhcp.php?if=opt7")
elemento = driver.find_elements(By.TAG_NAME,"td")
valores = []
valores_filtro=[]
sublistas  = []

for tag in elemento:
    valor = tag.text
    valores.append(valor)



valores_filtro= list(filter(lambda x: x !="",valores))

for i in range(0, len(valores_filtro), 5):
    sublista = valores_filtro[i:i+5]
    sublistas.append(sublista)


data = {
    'Valor': sublistas,    
}


for i in range(len(sublistas)):
    sheet2.cell(row=x+2, column=1).value = str(sublistas[i][0])
    sheet2.cell(row=x+2, column=2).value = str(sublistas[i][1])
    sheet2.cell(row=x+2, column=3).value = str(sublistas[i][2])
    sheet2.cell(row=x+2, column=4).value = str(sublistas[i][3])
    sheet2.cell(row=x+2, column=5).value = str(sublistas[i][4])
    x+=1

print(x)

workbook.save('c:\\config\\Firewall.xlsx')


# Fecha o navegador
driver.quit()