import openpyxl
import os
import subprocess
import csv
import datetime
#Peso
peso = 0.400       

# Obtém o caminho do diretório atual
current_directory = os.getcwd()

# Cria um novo arquivo Excel
workbook = openpyxl.Workbook()
sheet = workbook.active

# Mescla as células 
sheet.merge_cells('B1:Q1')
sheet.merge_cells('B2:B4')
sheet.merge_cells('B5:C5')
sheet.merge_cells('C2:L2')
sheet.merge_cells('C3:L3')
sheet.merge_cells('C4:L4')
sheet.merge_cells('E5:L5')
sheet.merge_cells('E6:L6')
sheet.merge_cells('Q3:Q5')




# Define o valor da célula mesclada
sheet['B1'] = 'CUSTO DE AMOSTRA'
sheet['B2'] = 'Lacres>>'
sheet['B5'] = 'Descrição>>>'
sheet['B6'] = '%'
sheet['C6'] = 'SKU.'
sheet['E6'] = 'Produtos'
sheet['Q2'] = 'Peso Unitário'
sheet['Q3'] = '0,400 Kg'
sheet['Q6'] = 'VALOR R$'


# Define o estilo dos textos
font = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=16)
font2 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=12,underline='single')
font3 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=10,underline='single')
font4 = openpyxl.styles.Font(name='Arial',bold=True, size=11)
font5 = openpyxl.styles.Font(name='Arial',bold=True, italic=True, size=12)
font6 = openpyxl.styles.Font(name='Arial',bold=True, size=10)
font7 = openpyxl.styles.Font(name='Arial',bold=True,  size=16)

#Define o alinhamento e aplica as fontes
sheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center')
sheet['B2'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['B6'].alignment = openpyxl.styles.Alignment(horizontal='center')
sheet['C6'].alignment = openpyxl.styles.Alignment(horizontal='center')
sheet['E6'].alignment = openpyxl.styles.Alignment(horizontal='center')
sheet['Q3'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
sheet['Q6'].alignment = openpyxl.styles.Alignment(horizontal='center')
sheet['B1'].font = font
sheet['B2'].font = font2
sheet['B5'].font = font5
sheet['B6'].font = font5
sheet['C6'].font = font5
sheet['E6'].font = font5
sheet['Q2'].font = font3
sheet['Q3'].font = font4
sheet['Q6'].font = font5


# Define a altura da linha 1
sheet.row_dimensions[1].height = 24.75
sheet.row_dimensions[6].height = 18.75
sheet.column_dimensions['B'].width = 10.14
sheet.column_dimensions['C'].width = 9.14
sheet.column_dimensions['Q'].width = 13.14
sheet.column_dimensions['D'].hidden = True
sheet.column_dimensions['E'].width = 6.86
sheet.column_dimensions['F'].width = 6.86
sheet.column_dimensions['G'].width = 6.86
sheet.column_dimensions['H'].width = 6.86
sheet.column_dimensions['I'].width = 6.86
sheet.column_dimensions['J'].width = 6.86
sheet.column_dimensions['K'].width = 6.86
sheet.column_dimensions['L'].width = 6.86

#Define as Bordas 
for col in range(13, 17):  # 13 = coluna M, 16 = coluna P
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True

all = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='medium'),
    right=openpyxl.styles.Side(border_style='medium'),
    top=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='medium')
)
all_thin = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='thin'),
    right=openpyxl.styles.Side(border_style='thin'),
    top=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='thin')
)

all_noright = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='medium'),
    top=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='medium')
)

all_notop = openpyxl.styles.Border(
    left=openpyxl.styles.Side(border_style='medium'),
    right=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='medium')
)

all_bottom = openpyxl.styles.Border(
    bottom=openpyxl.styles.Side(border_style='medium')
)

all_right = openpyxl.styles.Border(
    right=openpyxl.styles.Side(border_style='medium')
)

all_right_bottom = openpyxl.styles.Border(
    right=openpyxl.styles.Side(border_style='medium'),
    bottom=openpyxl.styles.Side(border_style='medium')    
)

for row in range(1, 2):  
    for col in range(2, 18): 
        sheet.cell(row=row, column=col).border = all

for row in range(2, 5):  
    sheet.cell(row=row,column=2).border = all_noright
    sheet.cell(row=row,column=17).border = all_right
    sheet.cell(row=row,column=12).border = all_right

for column in range(3, 13):  
    sheet.cell(row=4,column=column).border = all_bottom

for column in range(5, 13):  
    sheet.cell(row=5,column=column).border = all

for column in range(2, 4):  
    sheet.cell(row=5,column=column).border = all

for column in range(2, 18):  
    sheet.cell(row=6,column=column).border = all


sheet.cell(row=4,column=12).border = all_right_bottom
sheet.cell(row=5,column=17).border = all_notop
sheet.sheet_view.showGridLines = False


csv_file_path = 'X:\\TI\\Giovane\\Amostra\\CUSTO.CSV'  # Substitua 'arquivo.csv' pelo nome do seu arquivo CSV
with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    next(reader)  # Pula a primeira linha (cabeçalho)
    subtotal = 0
    row_num2 = 7
    for row in reader:
        if any(cell.strip() for cell in row):
            if row[0] == "MP":  
                DescricaoSDP = row[3] 
                Sku = row[4]
                DescricaoPRD = row[5]
                Qtde = row[6] 
                Valor = row[7]    
                sheet.cell(row=5, column=5).value = DescricaoSDP
                sheet.cell(row=5, column=5).alignment =openpyxl.styles.Alignment(horizontal='center') 
                sheet.cell(row=row_num2, column=2).value = float(Qtde.replace(',', '.'))            
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12)
                sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center') 
                sheet.cell(row=row_num2, column=3).value = int(Sku)
                sheet.cell(row=row_num2, column=3).alignment =openpyxl.styles.Alignment(horizontal='center')
                sheet.cell(row=row_num2, column=5).value = DescricaoPRD 
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='center')    
                sheet.cell(row=row_num2, column=17).value = peso * float(Valor.replace(',', '.')) * (float(Qtde.replace(',', '.'))) 
                sheet.cell(row=row_num2, column=17).alignment =openpyxl.styles.Alignment(horizontal='center')
                subtotal = subtotal + (peso * float(Valor.replace(',', '.')) * (float(Qtde.replace(',', '.')))) 
                sheet.cell(row=row_num2, column=2).border = all_thin
                sheet.cell(row=row_num2, column=3).border = all_thin            
                sheet.cell(row=row_num2, column=17).border = all_thin

                for col in range(5, 13):
                    sheet.cell(row=row_num2, column=col).border = all_thin

                row_num2 += 1

                for row_num3 in range(row_num2, row_num2 + 2):
                    for col in range(2, 13):
                        sheet.merge_cells(start_row=row_num3,end_row=row_num3,start_column=5,end_column=12)
                        sheet.cell(row=row_num3, column=col).border = all_thin
                    sheet.cell(row=row_num3, column=17).border = all_thin        
    row_num2 += 2       
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=12)    
    sheet.cell(row=row_num2, column=2).value = "MANUAIS / LAVAGEM"
    sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center')
    sheet.cell(row=row_num2, column=2).font = font
    for col in range(2, 18):
        sheet.cell(row=row_num2, column=col).border = all_thin
    row_num2 += 1
with open(csv_file_path, 'r', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=';')
    next(reader)  # Pula a primeira linha (cabeçalho)
    subtotal = 0
    #row_num2 = 21
    for row in reader:
        if any(cell.strip() for cell in row):
            if row[0] == "SV": 
                sheet.cell(row=row_num2, column=3).value = int(row[4]) 
                sheet.cell(row=row_num2, column=3).alignment =openpyxl.styles.Alignment(horizontal='center')
                sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12)
                sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='center') 
                sheet.cell(row=row_num2, column=5).value = row[5]                 
                for col in range(2, 18):
                    sheet.cell(row=row_num2, column=col).border = all_thin 
                row_num2 += 1
                for row_num3 in range(row_num2, row_num2 + 3):
                    for col in range(2, 13):
                        sheet.merge_cells(start_row=row_num3,end_row=row_num3,start_column=5,end_column=12)
                        sheet.cell(row=row_num3, column=col).border = all_thin
                        sheet.cell(row=row_num3, column=17).border = all_thin            
    row_num2 += 2        
    sheet.row_dimensions[row_num2].height = 5.0
    for col in range(2,18):
                    sheet.cell(row=row_num2, column=col).border = all_thin
    row_num2 += 1
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
    data = datetime.date.today()
    data_formatada = data.strftime('%d/%m/%Y')
    sheet.row_dimensions[row_num2].height = 24
    sheet.cell(row=row_num2, column=2).value = data_formatada
    sheet.cell(row=row_num2, column=2).font = font6
    sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
    sheet.row_dimensions[row_num2].height = 24
    for col in range(2,18):
        sheet.cell(row=row_num2, column=col).border = all_thin 
        sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
        sheet.cell(row=row_num2, column=5).value = "CUSTO PRODUTOS + 20%"
        sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
        sheet.cell(row=row_num2, column=5).font = font6 
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
    sheet.cell(row=row_num2, column=2).value = "CALCULADO POR:"
    sheet.cell(row=row_num2, column=2).font = font6
    sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
    for col in range(2,18):
        sheet.cell(row=row_num2, column=col).border = all_thin 
        sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
        sheet.cell(row=row_num2, column=5).value = "CUSTO MANUAIS/LAVAGEM"
        sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
        sheet.cell(row=row_num2, column=5).font = font6
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=2,end_column=3)
    sheet.cell(row=row_num2, column=2).value = " "
    sheet.cell(row=row_num2, column=2).font = font6
    sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center')
    for col in range(2,18):
        sheet.cell(row=row_num2, column=col).border = all_thin 
        sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
        sheet.cell(row=row_num2, column=5).value = "CUSTO SUB-TOTAL >>>"
        sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center') 
        sheet.cell(row=row_num2, column=5).font = font6
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2+4,start_column=2,end_column=3)
    sheet.cell(row=row_num2, column=2).value = "ATENÇÃO NOVOS PREÇOS "
    sheet.cell(row=row_num2, column=2).alignment =openpyxl.styles.Alignment(horizontal='center',vertical='center',wrap_text=True) 
    sheet.cell(row=row_num2, column=2).font = font7
    for row1 in range(row_num2,row_num2+5):
        for col in range(2,18):
            sheet.cell(row=row1, column=col).border = all                
            sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
            sheet.cell(row=row_num2, column=5).value = "TRANSPORTE"
            sheet.cell(row=row_num2, column=5).font = font6
            sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
        for col in range(2,18):
            sheet.cell(row=row1, column=col).border = all  
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
    sheet.cell(row=row_num2, column=5).value = "COMISSÃO"
    sheet.cell(row=row_num2, column=5).font = font6
    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
    for col in range(2,18):
        sheet.cell(row=row1, column=col).border = all  
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
    sheet.cell(row=row_num2, column=5).value = "TAMANHO DAS PEÇAS"
    sheet.cell(row=row_num2, column=5).font = font6
    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
    for col in range(2,18):
        sheet.cell(row=row1, column=col).border = all  
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
    sheet.cell(row=row_num2, column=5).value = "PREÇO SUGERIDO PARA TERCEIROS (TOTAL + 20%)"
    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')
    sheet.cell(row=row_num2, column=5).font = font6
    for col in range(2,18):
        sheet.cell(row=row1, column=col).border = all   
    row_num2 += 1
    sheet.row_dimensions[row_num2].height = 24
    sheet.merge_cells(start_row=row_num2,end_row=row_num2,start_column=5,end_column=12) 
    sheet.cell(row=row_num2, column=5).value = "CUSTO TOTAL >>>"
    sheet.cell(row=row_num2, column=5).alignment =openpyxl.styles.Alignment(horizontal='right',vertical='center')  
    sheet.cell(row=row_num2, column=5).font = font6              
    for col in range(2,18):
        sheet.cell(row=row1, column=col).border = all
    
    

            
                                   
  


                
                



workbook.save(os.path.join(current_directory, 'Custo.xlsx'))
print('Arquivo Excel criado com sucesso!')